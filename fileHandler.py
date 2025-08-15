import os
import re
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

# -------------------------------
# Helpers
# -------------------------------
_space_re = re.compile(r"[\u0020\u00A0\u2000-\u200B\u202F\u205F]")


def to_number(val):
    """Преобразует строки формата "2 335,99"/"1\u202F234,50"/"-7\u00A0710,11" -> float.
    Поддержка отрицательных в скобках: (1\u00A0234,50) -> -1234.5
    Поддержка юникод-минуса (\u2212).
    Если не число — возвращает исходное значение.
    """
    if isinstance(val, (float, int)):
        return float(val)
    if val is None:
        return np.nan

    s = str(val).strip()
    if s == "" or s.lower() in {"nan", "none", "null"}:
        return np.nan

    # Юникод-минус -> обычный минус
    s = s.replace("\u2212", "-")

    # Отрицательные в скобках: (123,45) -> -123,45
    is_neg = False
    if s.startswith("(") and s.endswith(")"):
        is_neg = True
        s = s[1:-1].strip()

    # Удаляем все виды пробелов/разделителей тысяч
    s_nospace = _space_re.sub("", s)

    # Запятая как десятичный разделитель -> точка
    s_dot = s_nospace.replace(",", ".")

    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", s_dot or ""):
        try:
            num = float(s_dot)
            return -num if is_neg else num
        except ValueError:
            return val
    else:
        return val


def _read_excel_auto(file_path: str) -> pd.DataFrame:
    """Читает ТОЛЬКО .xls (это осознанная логика бота) и приводит всё к str для единообразной чистки.
    - .xls => engine="xlrd" (нужен xlrd==1.2.0). Если его нет — подсказываем пользователю.
    - Любые другие расширения (.xlsx/.xlsm и т.д.) — вызывают понятную ошибку.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        try:
            return pd.read_excel(file_path, sheet_name=0, dtype=str, engine="xlrd")
        except Exception as e:
            raise RuntimeError(
                "Не удалось прочитать .xls. Убедитесь, что установлен xlrd==1.2.0 (в xlrd>=2.0 поддержка .xls удалена).\n"
                "Команда: pip install 'xlrd==1.2.0'"
            ) from e
    # Нам намеренно НЕ нужны .xlsx/.xlsm и прочие форматы на входе
    raise RuntimeError(
        "Поддерживается только входной формат .xls. По задумке бота .xlsx/.xlsm не отправляются."
    )


# -------------------------------
# Основная функция для бота
# -------------------------------

def process_excel(
    file_path: str,
    output_path: str,
    allowed_deviation_percentage: float = 0.035,
) -> str:
    """
    Обрабатывает ОСВ-файл и сохраняет результат в output_path.

    :param file_path: путь к входному Excel (.xls)
    :param output_path: путь к выходному .xlsx
    :param allowed_deviation_percentage: допустимое отклонение от оборота (доля, например 0.035 для 3.5%)
    :return: output_path (для удобства)
    """
    # --- читаем входной файл ---
    df = _read_excel_auto(file_path)

    # --- переименуем "Unnamed" колонки (кроме самой первой) ---
    cols = list(df.columns)
    if cols:
        first = cols[0]
        new_cols = [first]
        for c in cols[1:]:
            name = "" if (c is None or str(c).startswith("Unnamed")) else c
            new_cols.append(name)
        df.columns = new_cols

    # --- удаляем строки, где ЛЮБАЯ ячейка содержит целые слова "Товар" или "кол-во" (без учета регистра) ---
    compiled = re.compile(r"\b(?:товар|кол-во)\b", flags=re.IGNORECASE)
    mask_has_words = df.apply(lambda s: s.astype(str).str.contains(compiled, na=False)).any(axis=1)

    # Защитим некоторые строки от удаления (по индексам pandas)
    protected_rows = [7, 8, 9]
    mask_to_drop = mask_has_words & ~df.index.isin(protected_rows)
    df = df.loc[~mask_to_drop].reset_index(drop=True)

    # --- конвертируем все ячейки в числа, где это возможно ---
    df = df.map(to_number)

    # --- Вписываем подписи в строку 7 (pandas) — это 9-я строка Excel ---
    needed_last_col_index = 26  # AA (0-based индекс)
    if df.shape[1] - 1 < needed_last_col_index:
        for _ in range(needed_last_col_index + 1 - df.shape[1]):
            df[f"Unnamed_{df.shape[1]}"] = pd.NA

    labels = [
        (7, 21, "Отклонение излишков"),              # V9
        (7, 22, "Отклонение недостач"),              # W9
        (7, 23, "Норма отклонения"),                 # X9
        (7, 24, "Кол-во превышения нормы"),          # Y9
        (7, 25, "Сумма превышения нормы излишков"),  # Z9
        (7, 26, "Сумма превышения нормы недостач"),  # AA9
    ]
    for r, c, val in labels:
        df.iat[r, c] = val

    # --- Индексы столбцов (pandas 0-based) ---
    A_IDX = 0
    J_IDX, K_IDX = 9, 10
    L_IDX, M_IDX = 11, 12
    N_IDX, O_IDX = 13, 14
    P_IDX, R_IDX = 15, 17
    T_IDX, U_IDX = 19, 20
    V_IDX, W_IDX, X_IDX, Y_IDX, Z_IDX, AA_IDX = 21, 22, 23, 24, 25, 26

    EPS = 1e-9

    def is_number(x) -> bool:
        return isinstance(x, (int, float)) and not (isinstance(x, float) and np.isnan(x))

    def nzf(x: object) -> float:
        try:
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return 0.0
            return float(x)
        except Exception:
            return 0.0

    pct_display = (f"{allowed_deviation_percentage * 100:.1f}".replace(".", ",") + "%")

    for i in range(df.shape[0]):
        if i == 7:  # пропускаем строку с подписями
            continue

        first_val = df.iat[i, A_IDX] if A_IDX < df.shape[1] else None
        if not is_number(first_val):
            continue  # считаем только строки, где первая колонка — число

        # исходные значения
        j = nzf(df.iat[i, J_IDX]) if J_IDX < df.shape[1] else 0.0
        l = nzf(df.iat[i, L_IDX]) if L_IDX < df.shape[1] else 0.0
        n = nzf(df.iat[i, N_IDX]) if N_IDX < df.shape[1] else 0.0
        p = nzf(df.iat[i, P_IDX]) if P_IDX < df.shape[1] else 0.0
        r_ = nzf(df.iat[i, R_IDX]) if R_IDX < df.shape[1] else 0.0
        t = nzf(df.iat[i, T_IDX]) if T_IDX < df.shape[1] else 0.0
        u = df.iat[i, U_IDX] if U_IDX < df.shape[1] else None
        f = nzf(df.iat[i, 5]) if 5 < df.shape[1] else 0.0  # F
        g = df.iat[i, 6] if 6 < df.shape[1] else None      # G
        h = nzf(df.iat[i, 7]) if 7 < df.shape[1] else 0.0  # H
        ii = df.iat[i, 8] if 8 < df.shape[1] else None     # I
        k = df.iat[i, K_IDX] if K_IDX < df.shape[1] else None
        m = df.iat[i, M_IDX] if M_IDX < df.shape[1] else None
        o = df.iat[i, O_IDX] if O_IDX < df.shape[1] else None

        # оборот и X
        turnover = abs(j) + abs(l) + abs(n)
        x_val = allowed_deviation_percentage * turnover
        df.iat[i, X_IDX] = x_val

        # V по P (излишки)
        if turnover <= EPS and abs(p) > EPS:
            df.iat[i, V_IDX] = "Излишек при нулевом обороте"
        elif abs(p) > x_val:
            df.iat[i, V_IDX] = f"Превышение {pct_display} от оборота"
        else:
            df.iat[i, V_IDX] = "Норма"

        # W по R (недостачи)
        if turnover <= EPS and abs(r_) > EPS:
            df.iat[i, W_IDX] = "Недостача при нулевом обороте"
        elif abs(r_) > x_val:
            df.iat[i, W_IDX] = f"Превышение {pct_display} от оборота"
        else:
            df.iat[i, W_IDX] = "Норма"

        # Y = MIN exceedance for P/R vs X, иначе "Норма"
        cond_p = abs(p) > x_val
        cond_r = abs(r_) > x_val
        if cond_p or cond_r:
            d_p = abs(p) - x_val if cond_p else float("inf")
            d_r = abs(r_) - x_val if cond_r else float("inf")
            df.iat[i, Y_IDX] = min(d_p, d_r)
        else:
            df.iat[i, Y_IDX] = "Норма"

        # Z = ЕСЛИ(|P|>X; (|P|-X) * tiered_ratio; "")
        z_val: Optional[float | str] = ""
        if abs(p) > x_val:
            ratio = None
            if abs(t) > EPS and is_number(u):
                ratio = abs(float(u)) / abs(t)
            elif abs(f) > EPS and is_number(g):
                ratio = float(g) / f
            elif abs(h) > EPS and is_number(ii):
                ratio = abs(float(ii)) / abs(h)
            elif abs(j) > EPS and is_number(k):
                ratio = abs(float(k)) / abs(j)
            elif abs(l) > EPS and is_number(m):
                ratio = abs(float(m)) / abs(l)
            elif abs(n) > EPS and is_number(o):
                ratio = abs(float(o)) / abs(n)

            if ratio is not None:
                z_val = (abs(p) - x_val) * ratio
            else:
                z_val = ""
        df.iat[i, Z_IDX] = z_val

        # AA = ЕСЛИ(|R|>X; (|R|-X) * tiered_ratio; "")
        aa_val: Optional[float | str] = ""
        if abs(r_) > x_val:
            ratio = None
            if abs(t) > EPS and is_number(u):
                ratio = abs(float(u)) / abs(t)
            elif abs(f) > EPS and is_number(g):
                ratio = float(g) / f
            elif abs(h) > EPS and is_number(ii):
                ratio = abs(float(ii)) / abs(h)
            elif abs(j) > EPS and is_number(k):
                ratio = abs(float(k)) / abs(j)
            elif abs(l) > EPS and is_number(m):
                ratio = abs(float(m)) / abs(l)
            elif abs(n) > EPS and is_number(o):
                ratio = abs(float(o)) / abs(n)

            if ratio is not None:
                aa_val = (abs(r_) - x_val) * ratio
            else:
                aa_val = ""
        df.iat[i, AA_IDX] = aa_val

    # --- Суммы по колонкам Z и AA и запись в строку "Итого" ---
    z_series = pd.to_numeric(df.iloc[:, Z_IDX], errors="coerce").fillna(0)
    aa_series = pd.to_numeric(df.iloc[:, AA_IDX], errors="coerce").fillna(0)

    def _row_has_itogo(row) -> bool:
        for v in row:
            if isinstance(v, str) and "итого" in v.casefold():
                return True
        return False

    itogo_mask = df.apply(_row_has_itogo, axis=1)
    if itogo_mask.any():
        itogo_idx = itogo_mask[itogo_mask].index[-1]  # последняя строка "Итого"
        # не включаем саму строку "Итого" в сумму
        sum_z = z_series[z_series.index != itogo_idx].sum()
        sum_aa = aa_series[aa_series.index != itogo_idx].sum()
        df.iat[itogo_idx, Z_IDX] = float(sum_z)
        df.iat[itogo_idx, AA_IDX] = float(sum_aa)

    # --- Сохраняем и форматируем ---
    # Важно: при header=True (по умолчанию) pandas-строка 7 окажется на 9-й строке Excel.
    header_written = True
    startrow = 0
    row_in_df = 7
    excel_label_row = startrow + (1 if header_written else 0) + row_in_df + 1  # = 9

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=startrow)
        ws = writer.book.active

        # Ширина колонок V..AA (22..27) и перенос текста
        for col_idx in range(22, 28):  # 22=V, 27=AA
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 35
            ws[f"{col_letter}{excel_label_row}"].alignment = Alignment(wrap_text=True, vertical="top")

        # Цвета для V9 и W9
        fill_v = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")  # V9
        fill_w = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")  # W9
        ws[f"V{excel_label_row}"].fill = fill_v
        ws[f"W{excel_label_row}"].fill = fill_w

        # Закраска всей строки, если слово "Превышение" встречается в V или W
        fill_v_row = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")  # для V
        fill_w_row = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")  # для W

        for row_idx in range(2, ws.max_row + 1):  # начиная со 2-й строки (после заголовка df)
            v_val = str(ws[f"V{row_idx}"].value or "").lower()
            w_val = str(ws[f"W{row_idx}"].value or "").lower()
            has_prev_v = "превышение" in v_val or "излишек" in v_val
            has_prev_w = "превышение" in w_val or "недостача" in w_val
            if has_prev_v or has_prev_w:
                row_fill = fill_w_row if has_prev_w else fill_v_row
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill

        # Тонкие границы для всех ячеек с данными
        thin = Side(style="thin", color="FF000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

        # --- Жирным строки, где встречается "Итого", "Товар" или "Кол-во" ---
        # Нормализуем типографские дефисы к обычному "-" (кол-во/кол–во и т.п.)
        _dash_map = str.maketrans({
            "\u2011": "-",  # non-breaking hyphen
            "\u2013": "-",  # en dash
            "\u2014": "-",  # em dash
            "\u2212": "-",  # minus sign
        })

        pattern = re.compile(r"\b(итого|товар|кол-во)\b", flags=re.IGNORECASE)

        for row_idx in range(2, ws.max_row + 1):  # со 2-й строки, чтобы не трогать заголовок DF
            make_bold = False
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, str):
                    s = val.translate(_dash_map)
                    if pattern.search(s):
                        make_bold = True
                        break
            if make_bold:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    return output_path


# --------------------------------------
# CLI для локальной проверки (не обязателен для бота)
# --------------------------------------
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Обработка ОСВ: подсветка отклонений")
    parser.add_argument("input", help="Путь к входному Excel (.xls)")
    parser.add_argument("output", help="Путь к выходному .xlsx")
    parser.add_argument("--pct", type=float, default=3.5, help="Допустимый процент (по умолчанию 3.5)")
    args = parser.parse_args()

    pct_fraction = args.pct / 100.0
    result_path = process_excel(args.input, args.output, allowed_deviation_percentage=pct_fraction)
    print(f"Готово: {result_path}")
